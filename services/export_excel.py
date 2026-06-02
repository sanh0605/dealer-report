import io
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Styles ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
GROUP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GROUP_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
NEGATIVE_FONT = Font(size=11, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="medium"), bottom=Side(style="medium")
)

REGION_ORDER = [
    "MB1+MB2", "MB1", "MB2", "MB3+MB4", "MB3", "MB4",
    "MN1+MN3", "MN1", "MN3", "MN2", "MN4+MN5+MN6", "MN4", "MN5", "MN6",
    "Siêu thị", "Grand Total"
]

PRODUCT_GROUP_LABELS = {
    "Xe OEM": "XE OEM",
    "Xe Java": "XE JAVA",
    "Xe Giant": "XE GIANT",
    "Maxxis": "MAXXIS",
    "Phụ kiện": "PHỤ KIỆN",
    "Khác": "KHÁC",
}
PRODUCT_GROUP_ORDER = ["Xe OEM", "Xe Java", "Xe Giant", "Maxxis", "Phụ kiện", "Khác"]


def _apply_cell_style(cell, font=None, fill=None, border=None, alignment=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format


def _write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=HEADER_BORDER,
                          alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))


def _write_data_cell(ws, row, col, value, fmt=None, is_negative=False, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    use_font = font or (NEGATIVE_FONT if is_negative else NORMAL_FONT)
    _apply_cell_style(cell, font=use_font, fill=fill, border=THIN_BORDER)
    if fmt:
        cell.number_format = fmt


def _auto_width(ws, min_w=12, max_w=30):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_w), max_w)
            ws.column_dimensions[col_letter].width = best


def _get_ytd_data(main_data, reference_date):
    ref = pd.Timestamp(reference_date)
    return main_data[
        (main_data["date_transfer"].dt.year == ref.year) &
        (main_data["date_transfer"] <= ref)
    ]


def _get_month_range(reference_date):
    """Return list of (month_number, month_start_datetime) from Jan to selected month."""
    months = []
    for m in range(1, reference_date.month + 1):
        months.append((m, datetime(reference_date.year, m, 1)))
    return months


def _calc_regional_stats(data, targets_df=None):
    """Calculate stats by region_group for a given dataset."""
    # Exclude Sieu thi from regional groups to avoid double counting
    non_st_data = data[data["region"] != "Siêu thị"]

    stats = {}
    for rg in REGION_ORDER:
        if rg == "Grand Total":
            rg_data = data
        elif rg == "Siêu thị":
            rg_data = data[data["region"] == "Siêu thị"]
        else:
            rg_data = non_st_data[non_st_data["region_group"] == rg]

        if rg_data.empty and rg != "Grand Total":
            continue

        revenue = rg_data["sales_revenue"].sum()
        cost = rg_data["cost_of_goods"].sum()
        volume = rg_data["sales_volume"].sum()
        lng = revenue - cost
        lng_pct = lng / revenue if revenue != 0 else 0

        target = 0
        if targets_df is not None and not targets_df.empty and rg in targets_df["region_group"].values:
            target_row = targets_df[targets_df["region_group"] == rg]
            target = target_row["target_revenue"].sum()

        target_pct = revenue / target if target > 0 else 0

        stats[rg] = {
            "revenue": revenue,
            "cost": cost,
            "volume": volume,
            "lng": lng,
            "lng_pct": lng_pct,
            "target": target,
            "target_pct": target_pct,
        }
    return stats


def _build_tong_quan(ws, ytd_data, main_data, reference_date, targets_df, ar_df):
    # --- Table 1: Ke hoach 2026 ---
    r = 1
    ws.cell(row=r, column=1, value="KẾ HOẠCH 2026").font = Font(bold=True, size=14)
    r = 2
    _write_header_row(ws, r, ["Chỉ số", "Kênh áp dụng", "Target", "Kết Quả", "Tiến độ", "Deadline"])

    dl_data = ytd_data[ytd_data["region"] != "Siêu thị"]
    st_data = ytd_data[ytd_data["region"] == "Siêu thị"]

    kpi_rows = [
        ("Doanh số", "Đại lý", dl_data["sales_revenue"].sum()),
        ("Doanh số", "Siêu thị", st_data["sales_revenue"].sum()),
        ("Số lượng xe bán ra", "Đại lý", int(dl_data[dl_data["category"].str.upper() == "BIKES"]["sales_volume"].sum())),
        ("Số lượng xe bán ra", "Siêu thị", int(st_data[st_data["category"].str.upper() == "BIKES"]["sales_volume"].sum())),
    ]

    for i, (chi_so, kenh, ket_qua) in enumerate(kpi_rows):
        r = 3 + i
        target_val = 0
        if targets_df is not None and not targets_df.empty:
            t = targets_df[(targets_df["kpi"] == chi_so) & (targets_df["channel"] == kenh)]
            if not t.empty:
                target_val = t["target_value"].iloc[0]
        tien_do = ket_qua / target_val if target_val > 0 else 0

        for col, val in enumerate([chi_so, kenh, target_val, ket_qua, tien_do, datetime(reference_date.year, 12, 31)], 1):
            fmt = None
            if col in (3, 4):
                fmt = '#,##0'
            elif col == 5:
                fmt = '0.0%'
            _write_data_cell(ws, r, col, val, fmt=fmt)

    # --- Table 2: Thang XX/YYYY ---
    r = 14
    month_label = f"Tháng {reference_date.month:02d}/{reference_date.year}"
    ws.cell(row=r, column=1, value=month_label).font = Font(bold=True, size=14)

    months = _get_month_range(reference_date)
    quarterly_label = f"% LNG {reference_date.year}"
    headers2 = ["Khu Vực", "Doanh Số", "Target", "% Target", "LNG", "% LNG"]
    for m_num, m_start in months:
        headers2.append(f"% LNG T{m_num:02d}")
    headers2.append("% LNG YTD")

    r = 15
    _write_header_row(ws, r, headers2)

    # Calculate monthly data from main_data (unfiltered by sidebar)
    ref = pd.Timestamp(reference_date)
    monthly_data = main_data[
        (main_data["date_transfer"].dt.month == ref.month) &
        (main_data["date_transfer"].dt.year == ref.year)
    ]
    stats_monthly = _calc_regional_stats(monthly_data, targets_df)
    stats_ytd = _calc_regional_stats(ytd_data, targets_df)

    r = 16
    for rg in REGION_ORDER:
        if rg not in stats_monthly and rg not in stats_ytd:
            continue
        sm = stats_monthly.get(rg, {"revenue": 0, "target": 0, "target_pct": 0, "lng": 0, "lng_pct": 0})
        sy = stats_ytd.get(rg, {"lng_pct": 0})

        _write_data_cell(ws, r, 1, rg)
        _write_data_cell(ws, r, 2, sm["revenue"], fmt='#,##0', is_negative=sm["revenue"] < 0)
        _write_data_cell(ws, r, 3, sm["target"], fmt='#,##0')
        _write_data_cell(ws, r, 4, sm["target_pct"], fmt='0.0%')
        _write_data_cell(ws, r, 5, sm["lng"], fmt='#,##0', is_negative=sm["lng"] < 0)
        _write_data_cell(ws, r, 6, sm["lng_pct"], fmt='0.0%')

        col = 7
        for m_num, m_start in months:
            m_data = ytd_data[
                (ytd_data["date_transfer"].dt.month == m_num) &
                (ytd_data["date_transfer"].dt.year == reference_date.year)
            ]
            m_non_st = m_data[m_data["region"] != "Siêu thị"]
            if rg == "Siêu thị":
                m_data = m_data[m_data["region"] == "Siêu thị"]
            elif rg == "Grand Total":
                m_data = m_data
            else:
                m_data = m_non_st[m_non_st["region_group"] == rg]
            m_rev = m_data["sales_revenue"].sum()
            m_cost = m_data["cost_of_goods"].sum()
            m_lng_pct = (m_rev - m_cost) / m_rev if m_rev != 0 else 0
            _write_data_cell(ws, r, col, m_lng_pct, fmt='0.0%')
            col += 1

        _write_data_cell(ws, r, col, sy["lng_pct"], fmt='0.0%')

        is_total = rg == "Grand Total"
        if is_total:
            for c in range(1, col + 1):
                _apply_cell_style(ws.cell(row=r, column=c), font=TOTAL_FONT, fill=TOTAL_FILL)
        r += 1

    # --- Table 3: Luy ke ---
    r += 2
    ytd_label = f"Luỹ kế từ 01/01/{reference_date.year} - {reference_date.strftime('%d/%m/%Y')}"
    ws.cell(row=r, column=1, value=ytd_label).font = Font(bold=True, size=14)
    r += 1
    _write_header_row(ws, r, ["Khu Vực", "Doanh số", "Doanh thu", "% Thu hồi", "Target", "% Target"])

    ytd_stats = _calc_regional_stats(ytd_data, targets_df)

    if ar_df is not None and not ar_df.empty and "order_id" in ytd_data.columns:
        ar_merged = ar_df.merge(ytd_data[["order_id"]].drop_duplicates(), on="order_id", how="inner")
        ar_merged["doanh_thu"] = ar_merged["Deduction Amout"].fillna(0) + ar_merged["Paid Amout"].fillna(0)
        ar_with_dealer = ar_merged.merge(ytd_data[["order_id", "region", "region_group"]], on="order_id", how="left")
        ar_by_region = ar_with_dealer.groupby("region_group")["doanh_thu"].sum().to_dict()
        ar_sieuthi = ar_with_dealer[ar_with_dealer["region"] == "Siêu thị"]["doanh_thu"].sum()
        ar_total = ar_with_dealer["doanh_thu"].sum()
    else:
        ar_by_region = {}
        ar_sieuthi = 0
        ar_total = 0

    r += 1
    for rg in REGION_ORDER:
        if rg not in ytd_stats:
            continue
        ys = ytd_stats[rg]
        doanh_thu = ar_by_region.get(rg, 0)
        if rg == "Siêu thị":
            doanh_thu = ar_sieuthi
        elif rg == "Grand Total":
            doanh_thu = ar_total
        thu_hoi = doanh_thu / ys["revenue"] if ys["revenue"] > 0 else 0

        _write_data_cell(ws, r, 1, rg)
        _write_data_cell(ws, r, 2, ys["revenue"], fmt='#,##0', is_negative=ys["revenue"] < 0)
        _write_data_cell(ws, r, 3, doanh_thu, fmt='#,##0', is_negative=doanh_thu < 0)
        _write_data_cell(ws, r, 4, thu_hoi, fmt='0.0%')
        _write_data_cell(ws, r, 5, ys["target"], fmt='#,##0')
        _write_data_cell(ws, r, 6, ys["target_pct"], fmt='0.0%')

        if rg == "Grand Total":
            for c in range(1, 7):
                _apply_cell_style(ws.cell(row=r, column=c), font=TOTAL_FONT, fill=TOTAL_FILL)
        r += 1

    _auto_width(ws)


def _build_nhom_san_pham(ws, ytd_data, reference_date):
    months = _get_month_range(reference_date)

    headers = ["NHÓM SẢN PHẨM"]
    for m_num, _ in months:
        headers.append(f"T{m_num:02d}")
    headers.append("Grand Total")

    _write_header_row(ws, 1, headers)

    r = 2
    for pg_key in PRODUCT_GROUP_ORDER:
        pg_label = PRODUCT_GROUP_LABELS.get(pg_key, pg_key)
        pg_data = ytd_data[ytd_data["product_group"] == pg_key]
        pg_total_rev = pg_data["sales_revenue"].sum()
        pg_total_cost = pg_data["cost_of_goods"].sum()
        pg_total_vol = pg_data["sales_volume"].sum()
        pg_total_lng_pct = (pg_total_rev - pg_total_cost) / pg_total_rev if pg_total_rev > 0 else 0

        # Product group label row
        cell = ws.cell(row=r, column=1, value=pg_label)
        _apply_cell_style(cell, font=GROUP_FONT, fill=GROUP_FILL, border=THIN_BORDER)
        r += 1

        for metric_label, metric_total, fmt in [
            ("Doanh số", pg_total_rev, '#,##0'),
            ("Số lượng", pg_total_vol, '#,##0'),
            ("Lợi Nhuận Gộp", pg_total_lng_pct, '0.0%'),
        ]:
            _write_data_cell(ws, r, 1, metric_label)
            col = 2
            for m_num, _ in months:
                m_data = pg_data[
                    (pg_data["date_transfer"].dt.month == m_num) &
                    (pg_data["date_transfer"].dt.year == reference_date.year)
                ]
                if metric_label == "Doanh số":
                    val = m_data["sales_revenue"].sum()
                elif metric_label == "Số lượng":
                    val = int(m_data["sales_volume"].sum())
                else:
                    m_rev = m_data["sales_revenue"].sum()
                    m_cost = m_data["cost_of_goods"].sum()
                    val = (m_rev - m_cost) / m_rev if m_rev > 0 else 0
                _write_data_cell(ws, r, col, val, fmt=fmt, is_negative=(isinstance(val, (int, float)) and val < 0))
                col += 1
            _write_data_cell(ws, r, col, metric_total, fmt=fmt, is_negative=(isinstance(metric_total, (int, float)) and metric_total < 0))
            r += 1

    _auto_width(ws)


def _build_top10_sheet(ws, ytd_data, reference_date, product_group=None, title=None):
    months = _get_month_range(reference_date)

    ws.cell(row=1, column=1, value=title or "TOP 10 ĐẠI LÝ THEO DOANH SỐ").font = Font(bold=True, size=14)

    headers = ["NHÓM SẢN PHẨM"]
    for m_num, _ in months:
        headers.append(f"T{m_num:02d}")
    headers.append("Grand Total")
    _write_header_row(ws, 2, headers)

    data = ytd_data
    if product_group:
        data = ytd_data[ytd_data["product_group"] == product_group]

    # Aggregate by dealer
    dealer_monthly = {}
    for _, row in data.iterrows():
        dname = row.get("dealer_name", "Unknown")
        m = row["date_transfer"].month
        if dname not in dealer_monthly:
            dealer_monthly[dname] = {"monthly": {}, "total_rev": 0, "total_vol": 0}
        dealer_monthly[dname]["monthly"][m] = dealer_monthly[dname]["monthly"].get(m, {"rev": 0, "vol": 0})
        dealer_monthly[dname]["monthly"][m]["rev"] += row["sales_revenue"]
        dealer_monthly[dname]["monthly"][m]["vol"] += row["sales_volume"]
        dealer_monthly[dname]["total_rev"] += row["sales_revenue"]
        dealer_monthly[dname]["total_vol"] += row["sales_volume"]

    top10 = sorted(dealer_monthly.items(), key=lambda x: x[1]["total_rev"], reverse=True)[:10]

    r = 3
    grand_total_monthly = {m: {"rev": 0, "vol": 0} for m, _ in months}
    grand_total_rev = 0
    grand_total_vol = 0

    for dname, ddata in top10:
        cell = ws.cell(row=r, column=1, value=dname)
        _apply_cell_style(cell, font=GROUP_FONT, fill=GROUP_FILL, border=THIN_BORDER)
        r += 1

        for metric, total_val, fmt in [
            ("rev", ddata["total_rev"], '#,##0'),
            ("vol", ddata["total_vol"], '#,##0'),
        ]:
            _write_data_cell(ws, r, 1, "Doanh số" if metric == "rev" else "Số lượng")
            col = 2
            for m_num, _ in months:
                val = ddata["monthly"].get(m_num, {}).get(metric, 0)
                _write_data_cell(ws, r, col, val, fmt=fmt, is_negative=(isinstance(val, (int, float)) and val < 0))
                grand_total_monthly[m_num][metric] += val
                col += 1
            _write_data_cell(ws, r, col, total_val, fmt=fmt, is_negative=(isinstance(total_val, (int, float)) and val < 0))
            if metric == "rev":
                grand_total_rev += total_val
            else:
                grand_total_vol += total_val
            r += 1

    # Grand Total row
    for metric, total_val, fmt in [
        ("rev", grand_total_rev, '#,##0'),
        ("vol", grand_total_vol, '#,##0'),
    ]:
        _write_data_cell(ws, r, 1, "Grand Total", font=TOTAL_FONT, fill=TOTAL_FILL)
        col = 2
        for m_num, _ in months:
            val = grand_total_monthly[m_num][metric]
            _write_data_cell(ws, r, col, val, fmt=fmt, is_negative=(isinstance(val, (int, float)) and val < 0),
                             font=TOTAL_FONT, fill=TOTAL_FILL)
            col += 1
        _write_data_cell(ws, r, col, total_val, fmt=fmt, is_negative=(isinstance(total_val, (int, float)) and val < 0),
                         font=TOTAL_FONT, fill=TOTAL_FILL)
        r += 1

    _auto_width(ws)


def _build_top10_overall(ws, ytd_data, reference_date):
    months = _get_month_range(reference_date)

    ws.cell(row=1, column=1, value="TOP 10 ĐẠI LÝ THEO DOANH SỐ").font = Font(bold=True, size=14)

    headers = ["ĐẠI LÝ"]
    for m_num, _ in months:
        headers.append(f"T{m_num:02d}")
    headers.append("Grand Total")
    _write_header_row(ws, 2, headers)

    # Aggregate by dealer
    dealer_monthly = {}
    for _, row in ytd_data.iterrows():
        dname = row.get("dealer_name", "Unknown")
        m = row["date_transfer"].month
        if dname not in dealer_monthly:
            dealer_monthly[dname] = {"monthly": {}, "total": 0}
        dealer_monthly[dname]["monthly"][m] = dealer_monthly[dname]["monthly"].get(m, 0) + row["sales_revenue"]
        dealer_monthly[dname]["total"] += row["sales_revenue"]

    top10 = sorted(dealer_monthly.items(), key=lambda x: x[1]["total"], reverse=True)[:10]

    r = 3
    grand_total_monthly = {m: 0 for m, _ in months}
    grand_total = 0

    for dname, ddata in top10:
        _write_data_cell(ws, r, 1, dname)
        col = 2
        for m_num, _ in months:
            val = ddata["monthly"].get(m_num, 0)
            _write_data_cell(ws, r, col, val, fmt='#,##0', is_negative=val < 0)
            grand_total_monthly[m_num] += val
            col += 1
        _write_data_cell(ws, r, col, ddata["total"], fmt='#,##0', is_negative=ddata["total"] < 0)
        grand_total += ddata["total"]
        r += 1

    # Grand Total
    _write_data_cell(ws, r, 1, "Grand Total", font=TOTAL_FONT, fill=TOTAL_FILL)
    col = 2
    for m_num, _ in months:
        val = grand_total_monthly[m_num]
        _write_data_cell(ws, r, col, val, fmt='#,##0', font=TOTAL_FONT, fill=TOTAL_FILL, is_negative=val < 0)
        col += 1
    _write_data_cell(ws, r, col, grand_total, fmt='#,##0', font=TOTAL_FONT, fill=TOTAL_FILL, is_negative=grand_total < 0)

    _auto_width(ws)


def _build_raw_data(ws, ytd_data):
    column_map = {
        "order_id": "Order Reference",
        "order_date": "Date Order",
        "date_transfer": "Date",
        "month_year": "Month",
        "channel_name": "Sale Channel",
        "dealer_name": "Customer",
        "region": "Dealer Region",
        "region_group": "Dealer Region Group",
        "item_id": "SKU",
        "brand": "Brand",
        "product_group": "Brand Group",
        "category": "Category",
        "subcategory": "Sub Category",
        "sales_revenue": "Revenue",
        "cost_of_goods": "applied_cost",
        "sales_volume": "Quantity",
    }

    # Calculate Gross Profit
    export_df = ytd_data.copy()
    export_df["gross_profit"] = export_df["sales_revenue"].fillna(0) - export_df["cost_of_goods"].fillna(0)

    # Select and rename columns
    available_cols = {k: v for k, v in column_map.items() if k in export_df.columns}
    export_df = export_df[list(available_cols.keys())].rename(columns=available_cols)

    # Add Gross Profit after applied_cost
    cols = list(export_df.columns)
    if "applied_cost" in cols:
        gp_idx = cols.index("applied_cost") + 1
        cols.insert(gp_idx, "Gross Profit")
        export_df = export_df.reindex(columns=cols)

    # Write headers
    for i, col_name in enumerate(export_df.columns, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=HEADER_BORDER)

    # Write data
    for r_idx, (_, row) in enumerate(export_df.iterrows(), 2):
        for c_idx, col_name in enumerate(export_df.columns, 1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            fmt = None
            if col_name in ("Revenue", "applied_cost", "Gross Profit"):
                fmt = '#,##0'
            elif col_name == "Quantity":
                fmt = '#,##0'
            is_neg = isinstance(val, (int, float)) and val is not None and val < 0
            _write_data_cell(ws, r_idx, c_idx, val, fmt=fmt, is_negative=is_neg)

    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_width(ws, min_w=10, max_w=25)


def generate_dashboard_excel(
    filtered_data: pd.DataFrame,
    previous_filtered_data: pd.DataFrame,
    main_data: pd.DataFrame,
    reference_date: datetime,
    period_label: str = "",
) -> bytes:
    from database.gsheets_db import read_sheet

    wb = openpyxl.Workbook()

    # Load supplementary data
    targets_df = read_sheet("sales_targets")
    ar_df = read_sheet("accounts_receivable")

    ytd_data = _get_ytd_data(main_data, reference_date)

    # Sheet 1: Tong Quan
    ws1 = wb.active
    ws1.title = "Tổng Quan"
    _build_tong_quan(ws1, ytd_data, main_data, reference_date, targets_df, ar_df)

    # Sheet 2: Nhom San Pham
    ws2 = wb.create_sheet("Nhóm Sản Phẩm")
    _build_nhom_san_pham(ws2, ytd_data, reference_date)

    # Sheet 3: Top 10 KH - DS (overall)
    ws3 = wb.create_sheet("Top 10 KH - DS")
    _build_top10_overall(ws3, ytd_data, reference_date)

    # Sheets 4-7: Top 10 per brand
    brand_sheets = [
        ("Xe OEM", "Top 10 KH - DS XE OEM", "TOP 10 ĐẠI LÝ THEO DOANH SỐ XE OEM"),
        ("Xe Java", "Top 10 KH - DS XE JAVA", "TOP 10 ĐẠI LÝ THEO DOANH SỐ XE JAVA"),
        ("Xe Giant", "Top 10 KH - DS XE GIANT", "TOP 10 ĐẠI LÝ THEO DOANH SỐ XE GIANT"),
        ("Maxxis", "Top 10 KH - DS MAXXIS", "TOP 10 ĐẠI LÝ THEO DOANH SỐ MAXXIS"),
    ]
    for pg_key, sheet_name, title in brand_sheets:
        ws = wb.create_sheet(sheet_name)
        _build_top10_sheet(ws, ytd_data, reference_date, product_group=pg_key, title=title)

    # Sheet 8: RAW DATA
    ws8 = wb.create_sheet("RAW DATA")
    _build_raw_data(ws8, ytd_data)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
